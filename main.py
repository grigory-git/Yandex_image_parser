
from funcs import *
import random
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
import os
import time
from random_user_agent.user_agent import UserAgent
from random_user_agent.params import SoftwareName, OperatingSystem
from openpyxl import *
from openpyxl.utils import column_index_from_string


#функция для поиска ячеек с заливкой в эксель-файле
#по умолчанию желтый цвет
def excel_find_colored_cells(file_path, sheet_name, column, hexColor ='FFFFFF00'):
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]
    colored_cells = []

    column_index = column_index_from_string(column)  # Преобразование буквенного обозначения в числовой индекс

    for row in sheet.iter_rows(min_row=1, max_col=column_index, max_row=sheet.max_row):
        for cell in row:
            if cell.fill.start_color.index == hexColor:  # Hex code for yellow
                colored_cells.append(cell.value)

    wb.close()
    wb = None
    return colored_cells

#функция добавляет строки в заданном столбце из экселевского листа
def add_excel_data_to_list(excel_sheet, first_row, last_row, list, column):
    for i in range(first_row, last_row+1):
        value = excel_sheet[column + str(i)].value
        if value not in ("", None):
            list.append(value)

software_names = [SoftwareName.FIREFOX.value]
operating_systems = [OperatingSystem.WINDOWS.value, OperatingSystem.LINUX.value]
user_agent_rotator = UserAgent(software_names=software_names, operating_systems=operating_systems, limit=100)

# Get list of user agents.
user_agents = user_agent_rotator.get_user_agents()

# Get Random User Agent String.
user_agent = user_agent_rotator.get_random_user_agent()
yellow_RGB = "FFFF00FF"
sheetName = 'Лист1'
mainpath = "E:\Quiz\\"
questions = mainpath+ "150 хитов\\" + "вопросы.xlsx"
wb = load_workbook(filename=questions,
                   read_only=True)
excel_sheet = wb[sheetName]
artist_column = "B"
artist_column2 = "H"
artists = []
first_row = 2
last_row = 161
add_excel_data_to_list(excel_sheet, first_row, last_row, artists, artist_column)
add_excel_data_to_list(excel_sheet, first_row, last_row, artists, artist_column2)
excel_sheet = None
wb.close()
wb = None
is_search_empty = True
element = None
First = True
options = webdriver.FirefoxOptions()
user_agent = user_agent_rotator.get_random_user_agent()
options.add_argument(f"--user-agent={user_agent}")
options.add_argument("--no-sandbox")
driver = webdriver.Firefox(options=options)
driver.get("https://yandex.ru/images")
time.sleep(random.randrange(8, 12, 1))
#element присваивается поисковая строка
element = driver.find_element(By.CLASS_NAME, "input__control")
#создаём перечиселние исполнителей и ищем, затем сохраняем картинку для каждого
for j, artist in enumerate(artists):
    print(artist)
    path = mainpath + "artists_photo\\" + artists[j] + ".jpg"
    if os.path.exists(path):
        continue
        print(f"файл изображения {artist} уже существует")
    #вводим исполнителя в поисковое поле, проверяем введено ли уже что-либо в поисковой строке
    if not is_search_empty:
        print("очищаем поисковую строку")
        element = driver.find_element("xpath", "//input[contains(@class,'HeaderForm-Input mini-suggest__input')]")
        for k in range(len(search) + 1):
            element.send_keys(Keys.BACKSPACE)
            is_search_empty = True
    #здесь пишем то, что будет перед названием группы/исполнителя в поисковой строке
    #например, "группа ", "музыка "
    added = ""
    search = added + artists[j]
    element.send_keys(search, Keys.RETURN)
    is_search_empty = False
    time.sleep(random.randrange(2, 4, 1))
    if First:
        try:
            clickbut = driver.find_element("xpath","//span[text()='Размер']")
            driver.execute_script("arguments[0].click();", clickbut)
            time.sleep(random.randrange(1, 3))
            clickbut = driver.find_element("xpath","//span[text()='Большие']")
            driver.execute_script("arguments[0].click();", clickbut)
            time.sleep(random.randrange(1, 3))
            clickbut = driver.find_element("xpath","//span[text()='Ориентация']")
            driver.execute_script("arguments[0].click();", clickbut)
            time.sleep(random.randrange(1, 3))
            clickbut = driver.find_element("xpath","//span[text()='Горизонтальные']")
            driver.execute_script("arguments[0].click();", clickbut)
            First = False
            time.sleep(random.randrange(1, 3))
        except:
            pass
    #проскролливаем страничку вниз, чтобы больше картинок загрузилось
    driver.execute_script("window.scrollTo(0,document.body.scrollHeight);")
    time.sleep(random.randrange(3, 6))
    images = driver.find_elements(By.XPATH,
                                  "//img[contains(@class,'ContentImage-Image')]")
    imgnum = 0
    time.sleep(random.randrange(3, 5))
    #пока не удастся сохранить файл необходимого размера продолжаем перебирать изображения
    while not (os.path.isfile(path) and conditions(path)):
        #открываем изображение
        driver.execute_script("arguments[0].click();", images[imgnum])
        time.sleep(random.randrange(1, 5))

            # Locate the 'a' element and retrieve the URL from the 'href' attribute
        #ждём пока изображение прогрузится
        image_link = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//a[contains(@class,'MMViewerButtons-OpenImage')]"))
        )
        url = image_link.get_attribute('href')

        # открываем изображение в новой вкладке
        driver.execute_script("window.open('');")
        url_opened = False
        # переходим в новую вкладку
        try:
            driver.switch_to.window(driver.window_handles[1])
            driver.get(url)
            url_opened = True
        except Exception as e:
            print("Не удалось перейти по ссылке. Продолжаем", e)
        time.sleep(random.randrange(2, 4))
        # ждём пока изображение загрузится
        if url_opened:
            success = True
            try:
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'img')))
            except Exception as e:
                success = False
                print(f"WebDriverWait error occurred: {e}")
            if success:
                img_element = driver.find_element(By.TAG_NAME, 'img')
                img_url = img_element.get_attribute('src')
                # сохраняем изображение
                save_image(img_url, path)
        while len(driver.window_handles) > 1:
            driver.close()
        driver.switch_to.window(driver.window_handles[0])
        imgnum += 1
        clickbut = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".ImagesViewer-Close"))
        )
        driver.execute_script("arguments[0].click();", clickbut)
        time.sleep(random.uniform(2, 4))

        # Wait for the new tab to open and switch to it
        #img_url = driver.find_element("xpath", "//img[contains(@class,'MMImage-Origin')]").get_attribute("src")



